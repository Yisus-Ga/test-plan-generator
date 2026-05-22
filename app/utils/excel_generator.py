"""
Utilidades para generar archivos Excel.
"""
import io
import re
import unicodedata
import pandas as pd
from typing import Tuple
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from app.utils.markdown_parser import parse_markdown_table_to_dict, extract_hu_id_from_markdown, extract_hu_title_from_markdown

# Ancho máximo de columna (caracteres). Contenido más largo se muestra con wrap text.
_MAX_COL_WIDTH = 60
_MIN_COL_WIDTH = 10

# Estilo del header
_HEADER_BG = "4472C4"   # Azul corporativo Excel
_HEADER_FG = "FFFFFF"


def clean_filename(name: str) -> str:
    """Limpia un nombre para que sea válido como nombre de archivo y compatible con headers HTTP (ASCII)."""
    # Normalizar unicode: descompone acentos y reemplaza caracteres como em-dash por equivalentes ASCII
    name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    return re.sub(r'[\\/*?:"<>|\s]+', "_", name).strip("_")


def _apply_excel_formatting(worksheet, df: pd.DataFrame) -> None:
    """
    Aplica al worksheet:
    - Auto-filter en la fila de encabezado.
    - Wrap text + alineación top en todas las celdas de datos.
    - Ancho de columna ajustado al contenido más largo (con tope en _MAX_COL_WIDTH).
    - Estilo visual en la fila de encabezado (fondo azul, texto blanco bold, centrado).
    """
    # Auto-filter sobre toda la tabla (incluye header)
    worksheet.auto_filter.ref = worksheet.dimensions

    # Calcular ancho óptimo por columna
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)

        # Longitud del header como punto de partida
        max_len = len(str(col_name))

        # Recorrer valores de la columna; si hay saltos de línea, tomar la línea más larga
        for value in df[col_name]:
            if value is not None:
                lines = str(value).splitlines()
                cell_max = max((len(line) for line in lines), default=0)
                max_len = max(max_len, cell_max)

        col_width = min(max(max_len + 2, _MIN_COL_WIDTH), _MAX_COL_WIDTH)
        worksheet.column_dimensions[col_letter].width = col_width

    # Formatear header (fila 1)
    header_fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    header_font = Font(color=_HEADER_FG, bold=True)
    header_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Wrap text + alineación top para todas las filas de datos
    data_alignment = Alignment(wrap_text=True, vertical="top")
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = data_alignment


def generate_excel_from_test_plan(
    hu_md: str,
    test_plan_md: str
) -> Tuple[io.BytesIO, str]:
    """
    Genera un archivo Excel en memoria a partir del Test Plan.

    Formato aplicado:
    - Wrap text en todas las celdas (sin truncamiento de contenido).
    - Ancho de columna proporcional al contenido más largo (máx. 60 caracteres).
    - Auto-filter nativo de Excel en la fila de encabezado.
    - Estilo visual en el header (fondo azul, texto blanco bold).

    Args:
        hu_md: Markdown de la Historia de Usuario
        test_plan_md: Markdown del Test Plan en español

    Returns:
        Tupla con (BytesIO del archivo, nombre del archivo)
    """
    hu_id = extract_hu_id_from_markdown(hu_md)
    hu_title = extract_hu_title_from_markdown(hu_md)
    filename = f"{hu_id}_{clean_filename(hu_title)}.xlsx"

    test_cases = parse_markdown_table_to_dict(test_plan_md)

    if test_cases:
        df = pd.DataFrame(test_cases)
        df.columns = ["Prioridad", "ID CP", "Título", "Precondición", "Validación Esperada", "Resultado Obtenido"]
    else:
        df = pd.DataFrame(columns=[
            "Prioridad", "ID CP", "Título", "Precondición",
            "Validación Esperada", "Resultado Obtenido"
        ])

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="TestPlan")
        _apply_excel_formatting(writer.sheets["TestPlan"], df)
    bio.seek(0)

    return bio, filename
